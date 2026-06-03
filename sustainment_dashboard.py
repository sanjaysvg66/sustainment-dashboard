import os
import json
import uuid
from pathlib import Path
from datetime import datetime, date

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── PAGE CONFIG ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Sustainment WIP Dashboard",
    page_icon="✈",
    layout="wide",
    initial_sidebar_state="expanded",
)

WIP_UPDATES_FILE = Path(r"\\svg-fs-03\ENGINEERING\Engineering\Sustainment\WIP\WIP List InProgress.xlsm")
WIP_FOLDER      = str(WIP_UPDATES_FILE.parent)
TRACKER_FILE    = Path(__file__).parent / "project_tracker_data.json"
UPDATES_SHEET   = "Dashboard Updates"

# Columns for the updates sheet
UPDATE_COLS = ["Timestamp", "Author", "Work Order", "Customer", "Part Number", "Update Type", "Priority", "Notes"]
UPDATE_TYPES  = ["Status Change", "Note / Comment", "Action Required", "Issue / Blocker", "Completed", "Customer Update", "Parts Update", "Other"]
UPDATE_COLORS = {
    "Status Change":   ("#1a4f8b", "badge-new"),
    "Note / Comment":  ("#374151", "badge"),
    "Action Required": ("#b45309", "badge-changed"),
    "Issue / Blocker": ("#991b1b", "badge-removed"),
    "Completed":       ("#166534", "badge-added"),
    "Customer Update": ("#1a4f8b", "badge-new"),
    "Parts Update":    ("#374151", "badge"),
    "Other":           ("#374151", "badge"),
}

# ─── GLOBAL CSS ───────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Rajdhani:wght@500;600;700&family=Share+Tech+Mono&display=swap');

:root {
    --bg-base:       #f0f2f5;
    --bg-panel:      #ffffff;
    --bg-surface:    #f7f8fa;
    --bg-card:       #ffffff;
    --accent:        #1a4f8b;
    --accent-light:  #e8f0fb;
    --accent-mid:    #2563ae;
    --amber:         #b45309;
    --amber-bg:      #fef3c7;
    --green:         #166534;
    --green-bg:      #dcfce7;
    --red:           #991b1b;
    --red-bg:        #fee2e2;
    --purple:        #5b21b6;
    --purple-bg:     #ede9fe;
    --text-main:     #000000;
    --text-secondary:#374151;
    --text-dim:      #6b7280;
    --border:        #d1d5db;
    --border-strong: #9ca3af;
    --shadow:        0 1px 3px rgba(0,0,0,0.1), 0 1px 2px rgba(0,0,0,0.06);
    --shadow-md:     0 4px 6px rgba(0,0,0,0.07), 0 2px 4px rgba(0,0,0,0.05);
}

html, body, [class*="css"], .stApp {
    font-family: Calibri, 'Segoe UI', sans-serif !important;
    background: var(--bg-base) !important;
    color: var(--text-main) !important;
}

#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 1.2rem 2rem 2rem; }

/* ── Masthead ── */
.masthead {
    display: flex; align-items: center; gap: 16px;
    padding: 16px 22px;
    background: var(--accent);
    border-radius: 6px;
    margin-bottom: 18px;
    box-shadow: var(--shadow-md);
}
.masthead-icon { font-size: 2rem; filter: brightness(10); }
.masthead-title {
    font-family: 'Rajdhani', sans-serif;
    font-size: 1.55rem; font-weight: 700;
    letter-spacing: 0.08em; text-transform: uppercase;
    color: #ffffff;
}
.masthead-sub {
    font-family: 'Share Tech Mono', monospace;
    font-size: 0.68rem; color: #a8c4e8;
    letter-spacing: 0.12em; text-transform: uppercase; margin-top: 2px;
}
.masthead-meta { margin-left: auto; text-align: right; }
.masthead-meta .ts { font-family: 'Share Tech Mono', monospace; font-size: 0.72rem; color: #a8c4e8; }
.masthead-meta .fn { font-size: 0.78rem; color: #d1e4f7; margin-top: 2px; }
.status-dot {
    display: inline-block; width: 8px; height: 8px; border-radius: 50%;
    background: #4ade80; box-shadow: 0 0 5px #4ade80;
    margin-right: 6px; animation: blink 2s infinite;
}
@keyframes blink { 0%,100%{opacity:1} 50%{opacity:.35} }

/* ── KPI Row ── */
.kpi-row { display: grid; grid-template-columns: repeat(4,1fr); gap: 12px; margin-bottom: 18px; }
.kpi-card {
    background: var(--bg-card); border: 1px solid var(--border);
    border-radius: 6px; padding: 14px 18px;
    box-shadow: var(--shadow); border-top: 3px solid transparent;
}
.kpi-card.blue  { border-top-color: var(--accent); }
.kpi-card.green { border-top-color: #16a34a; }
.kpi-card.amber { border-top-color: #d97706; }
.kpi-card.red   { border-top-color: #dc2626; }
.kpi-label {
    font-family: Calibri, sans-serif; font-size: 0.72rem;
    text-transform: uppercase; letter-spacing: 0.08em;
    color: var(--text-dim); margin-bottom: 6px; font-weight: 600;
}
.kpi-value {
    font-family: 'Rajdhani', sans-serif; font-size: 1.9rem;
    font-weight: 700; color: var(--text-main); line-height: 1;
}
.kpi-card.blue  .kpi-value { color: var(--accent); }
.kpi-card.green .kpi-value { color: #15803d; }
.kpi-card.amber .kpi-value { color: #b45309; }
.kpi-card.red   .kpi-value { color: #dc2626; }
.kpi-sub { font-size: 0.73rem; color: var(--text-dim); margin-top: 4px; }

/* ── Section heading ── */
.sec-head {
    font-family: 'Rajdhani', sans-serif;
    font-size: 1rem; font-weight: 700;
    text-transform: uppercase; letter-spacing: 0.1em;
    color: var(--accent); border-bottom: 2px solid var(--accent-light);
    padding-bottom: 6px; margin-bottom: 12px;
}

/* ── Panel ── */
.panel {
    background: var(--bg-panel); border: 1px solid var(--border);
    border-radius: 6px; padding: 18px 20px;
    box-shadow: var(--shadow); margin-bottom: 16px;
}

/* ── Table ── */
.table-wrap { overflow: auto; max-height: 700px; }
.excel-table {
    border-collapse: collapse;
    font-family: Calibri, Arial, sans-serif;
    font-size: 13px; color: #000000;
    background: white; width: max-content;
}
.excel-table td, .excel-table th {
    border: 1px solid #d1d5db; padding: 5px 10px;
    min-width: 100px; white-space: nowrap;
}
.excel-col-header {
    background: #1a4f8b !important; color: #ffffff !important;
    font-family: Calibri, sans-serif; font-size: 0.72rem;
    font-weight: 700; letter-spacing: 0.05em;
    text-align: center; position: sticky; top: 0; z-index: 2;
}
.excel-row-header {
    background: #f3f4f6 !important; color: #6b7280 !important;
    font-family: 'Share Tech Mono', monospace; font-size: 0.65rem;
    text-align: center; position: sticky; left: 0; z-index: 1;
    min-width: 44px !important; border-right: 2px solid #d1d5db !important;
}
.excel-corner {
    background: #1a4f8b !important; position: sticky;
    top: 0; left: 0; z-index: 3; min-width: 44px !important;
}
tr:hover td { background: #f0f6ff !important; }
.search-match { outline: 2px solid #d97706 !important; outline-offset:-2px; background: #fef9c3 !important; }

/* ── Diff table ── */
.diff-table { border-collapse: collapse; width: 100%; font-family: Calibri, sans-serif; font-size: 13px; }
.diff-table th {
    background: #1a4f8b; color: white; padding: 8px 12px;
    text-align: left; font-weight: 700; letter-spacing: 0.04em;
    position: sticky; top: 0; z-index: 1;
}
.diff-table td { padding: 7px 12px; border-bottom: 1px solid #e5e7eb; color: #000; }
.diff-table tr:hover td { background: #f0f6ff; }
.badge {
    display: inline-block; padding: 2px 8px; border-radius: 3px;
    font-size: 0.7rem; font-weight: 700; letter-spacing: 0.06em; text-transform: uppercase;
}
.badge-added   { background: var(--green-bg);  color: var(--green); }
.badge-removed { background: var(--red-bg);    color: var(--red); }
.badge-changed { background: var(--amber-bg);  color: var(--amber); }
.badge-new     { background: var(--accent-light); color: var(--accent); }

/* ── Project tracker ── */
.proj-card {
    background: white; border: 1px solid var(--border);
    border-radius: 6px; padding: 16px 18px;
    box-shadow: var(--shadow); margin-bottom: 10px;
    border-left: 4px solid transparent;
    transition: box-shadow 0.15s;
}
.proj-card:hover { box-shadow: var(--shadow-md); }
.proj-card.status-complete   { border-left-color: #16a34a; }
.proj-card.status-inprogress { border-left-color: var(--accent); }
.proj-card.status-onhold     { border-left-color: #d97706; }
.proj-card.status-notstarted { border-left-color: #9ca3af; }
.proj-card.status-cancelled  { border-left-color: #dc2626; }
.proj-title {
    font-family: Calibri, sans-serif; font-size: 1rem;
    font-weight: 700; color: #000; margin-bottom: 4px;
}
.proj-meta {
    font-size: 0.75rem; color: var(--text-dim); margin-bottom: 8px;
}
.proj-desc { font-size: 0.82rem; color: var(--text-secondary); margin-bottom: 10px; }
.prog-bar-wrap { background: #e5e7eb; border-radius: 99px; height: 7px; margin-bottom: 8px; }
.prog-bar-fill { height: 7px; border-radius: 99px; transition: width 0.3s; }
.prog-bar-fill.c-complete   { background: #16a34a; }
.prog-bar-fill.c-inprogress { background: var(--accent); }
.prog-bar-fill.c-onhold     { background: #d97706; }
.prog-bar-fill.c-notstarted { background: #9ca3af; }
.prog-bar-fill.c-cancelled  { background: #dc2626; }

/* ── Status / priority badges ── */
.st-complete   { background:#dcfce7; color:#166534; }
.st-inprogress { background:#dbeafe; color:#1e40af; }
.st-onhold     { background:#fef3c7; color:#92400e; }
.st-notstarted { background:#f3f4f6; color:#374151; }
.st-cancelled  { background:#fee2e2; color:#991b1b; }
.pr-critical   { background:#fee2e2; color:#991b1b; }
.pr-high       { background:#fef3c7; color:#92400e; }
.pr-medium     { background:#dbeafe; color:#1e40af; }
.pr-low        { background:#f3f4f6; color:#374151; }

/* ── Sidebar ── */
section[data-testid="stSidebar"] {
    background: #ffffff !important;
    border-right: 1px solid #d1d5db;
}
section[data-testid="stSidebar"] * { color: #000000 !important; }
section[data-testid="stSidebar"] label { font-size: 0.8rem !important; color: #374151 !important; }
.sidebar-sec {
    font-family: 'Rajdhani', sans-serif; font-size: 0.8rem; font-weight: 700;
    letter-spacing: 0.12em; text-transform: uppercase; color: #1a4f8b !important;
    border-bottom: 1px solid #d1d5db; padding-bottom: 4px; margin: 14px 0 8px;
}

/* ── Streamlit widget overrides ── */
.stSelectbox > div > div,
.stTextInput > div > div > input,
.stNumberInput > div > div > input,
.stTextArea textarea,
.stDateInput input {
    background: #f9fafb !important; border-color: #d1d5db !important;
    color: #000000 !important; font-family: Calibri, sans-serif !important;
    font-size: 0.9rem !important;
}
.stButton > button {
    background: var(--accent) !important; color: #ffffff !important;
    border: none !important; font-family: Calibri, sans-serif !important;
    font-weight: 700 !important; border-radius: 4px !important;
    letter-spacing: 0.04em !important;
}
.stButton > button:hover { background: var(--accent-mid) !important; }
div[data-testid="stTabs"] button {
    font-family: 'Rajdhani', sans-serif !important; font-weight: 600 !important;
    font-size: 0.95rem !important; letter-spacing: 0.06em !important;
    text-transform: uppercase !important; color: var(--text-secondary) !important;
}
div[data-testid="stTabs"] button[aria-selected="true"] {
    color: var(--accent) !important; border-bottom-color: var(--accent) !important;
}
.stExpander { background: white !important; border: 1px solid var(--border) !important; border-radius: 6px !important; }
label, p, li, span { color: #000000 !important; }
h1,h2,h3 { color: #000000 !important; }
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #f3f4f6; }
::-webkit-scrollbar-thumb { background: #d1d5db; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: #9ca3af; }

/* ── Filter panel ── */
.filter-panel {
    background: #ffffff; border: 1px solid #d1d5db;
    border-radius: 8px; padding: 0;
    margin-bottom: 14px; box-shadow: var(--shadow); overflow: hidden;
}
.filter-panel-header {
    background: linear-gradient(90deg, #1a4f8b 0%, #2563ae 100%);
    padding: 10px 16px; display: flex; align-items: center;
    justify-content: space-between; gap: 10px;
}
.filter-panel-title {
    font-family: 'Rajdhani', sans-serif; font-size: 0.88rem; font-weight: 700;
    text-transform: uppercase; letter-spacing: 0.1em; color: #ffffff;
}
.filter-panel-body { padding: 14px 16px 10px; background: #f8fafc; border-top: 1px solid #e2e8f0; }
.filter-active-bar {
    padding: 8px 16px; background: #eff6ff; border-top: 1px solid #bfdbfe;
    display: flex; align-items: center; gap: 6px; flex-wrap: wrap; min-height: 36px;
}
.filter-chip {
    display: inline-flex; align-items: center; gap: 4px;
    background: #1a4f8b; color: #ffffff !important;
    font-family: Calibri, sans-serif; font-size: 0.72rem; font-weight: 700;
    padding: 3px 10px; border-radius: 99px; white-space: nowrap;
}
.filter-chip-col { opacity: 0.75; font-weight: 400; }
.filter-chip-val { color: #ffffff !important; }
.filter-none-label { font-size: 0.75rem; color: #9ca3af; font-style: italic; }
.row-count-bar {
    background: #ffffff; border: 1px solid #d1d5db;
    border-radius: 6px; padding: 8px 16px;
    display: flex; align-items: center; gap: 10px;
    margin-bottom: 10px; box-shadow: var(--shadow);
}
.row-count-number { font-family: 'Rajdhani', sans-serif; font-size: 1.3rem; font-weight: 700; color: #1a4f8b; }
.row-count-label { font-size: 0.78rem; color: #6b7280; }
.row-count-divider { width: 1px; height: 28px; background: #e5e7eb; margin: 0 4px; }

/* ── Multiselect polish ── */
div[data-baseweb="select"] { border-radius: 6px !important; }
div[data-baseweb="select"] > div {
    background: #ffffff !important; border: 1.5px solid #d1d5db !important;
    border-radius: 6px !important; font-family: Calibri, sans-serif !important;
    font-size: 0.85rem !important; color: #000000 !important;
    min-height: 36px !important; transition: border-color 0.15s !important;
}
div[data-baseweb="select"] > div:hover { border-color: #1a4f8b !important; }
div[data-baseweb="select"] > div:focus-within {
    border-color: #1a4f8b !important;
    box-shadow: 0 0 0 3px rgba(26,79,139,0.12) !important;
}
span[data-baseweb="tag"] { background: #1a4f8b !important; border-radius: 99px !important; padding: 1px 8px !important; }
span[data-baseweb="tag"] span { color: #ffffff !important; font-size: 0.75rem !important; font-family: Calibri, sans-serif !important; }
span[data-baseweb="tag"] button { color: #ffffff !important; opacity: 0.8; }
ul[data-baseweb="menu"] {
    background: #ffffff !important; border: 1.5px solid #d1d5db !important;
    border-radius: 6px !important; box-shadow: var(--shadow-md) !important;
}
ul[data-baseweb="menu"] li { font-family: Calibri, sans-serif !important; font-size: 0.85rem !important; color: #000000 !important; padding: 6px 12px !important; }
ul[data-baseweb="menu"] li:hover { background: #eff6ff !important; color: #1a4f8b !important; }
.filter-col-label {
    font-family: Calibri, sans-serif; font-size: 0.72rem; font-weight: 700;
    text-transform: uppercase; letter-spacing: 0.06em; color: #374151; margin-bottom: 4px;
}

/* ── Update card ── */
.update-card {
    background: white; border: 1px solid #d1d5db; border-radius: 6px;
    padding: 14px 18px; margin-bottom: 10px;
    border-left: 4px solid #1a4f8b; box-shadow: var(--shadow);
}
.update-card:hover { box-shadow: var(--shadow-md); }
</style>
""", unsafe_allow_html=True)


# ─── HELPERS ──────────────────────────────────────────────────────────────────
def get_excel_files(folder):
    try:
        files = sorted(
            [f for f in Path(folder).glob("*")
             if f.suffix.lower() in [".xlsx", ".xlsm"] and not f.name.startswith("~$")],
            key=os.path.getmtime, reverse=True
        )
        return files
    except Exception:
        return []

def latest_excel(folder):
    files = get_excel_files(folder)
    return files[0] if files else None

def color_to_hex(color):
    try:
        if color and color.type == "rgb" and color.rgb:
            raw = color.rgb[-6:]
            if raw.upper() not in ["000000", "FFFFFF", "ffffff"]:
                return "#" + raw
    except Exception:
        pass
    return None

def cell_has_style(cell):
    return (
        cell.value not in [None, ""] or cell.has_style
        or (cell.fill and cell.fill.fill_type not in [None, "none"])
        or cell.font.bold
        or cell.border.left.style is not None
    )

def detect_real_range(ws):
    max_row, max_col = 1, 1
    for row in ws.iter_rows():
        for cell in row:
            if cell_has_style(cell):
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    for rng in ws.merged_cells.ranges:
        b = rng.bounds
        max_row = max(max_row, b[3])
        max_col = max(max_col, b[2])
    return max_row, max_col

def cell_style(cell):
    styles = ["background-color:#ffffff", "color:#000000"]
    try:
        if cell.fill and cell.fill.fill_type == "solid":
            bg = color_to_hex(cell.fill.fgColor)
            if bg:
                styles[0] = f"background-color:{bg}"
    except Exception:
        pass
    try:
        if cell.font:
            fg = color_to_hex(cell.font.color)
            if fg:
                styles[1] = f"color:{fg}"
            if cell.font.bold:   styles.append("font-weight:bold")
            if cell.font.italic: styles.append("font-style:italic")
            if cell.font.sz:     styles.append(f"font-size:{cell.font.sz}px")
    except Exception:
        pass
    try:
        if cell.alignment:
            if cell.alignment.horizontal:
                styles.append(f"text-align:{cell.alignment.horizontal}")
            if cell.alignment.vertical:
                styles.append(f"vertical-align:{cell.alignment.vertical}")
    except Exception:
        pass
    return ";".join(styles)

def merged_map(ws):
    merged, skip = {}, set()
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = rng.bounds
        merged[(min_row, min_col)] = {"rowspan": max_row-min_row+1, "colspan": max_col-min_col+1}
        for r in range(min_row, max_row+1):
            for c in range(min_col, max_col+1):
                if not (r == min_row and c == min_col):
                    skip.add((r, c))
    return merged, skip

def collect_kpis(ws, max_row, max_col):
    non_empty, numeric_vals, bold_count, colored = 0, [], 0, 0
    for r in range(1, max_row+1):
        for c in range(1, max_col+1):
            cell = ws.cell(row=r, column=c)
            if cell.value not in [None, ""]:
                non_empty += 1
                if isinstance(cell.value, (int, float)):
                    numeric_vals.append(cell.value)
            try:
                if cell.font and cell.font.bold: bold_count += 1
                if cell.fill and cell.fill.fill_type == "solid": colored += 1
            except Exception:
                pass
    return {
        "total_rows": max_row, "total_cols": max_col,
        "non_empty": non_empty, "numeric_count": len(numeric_vals),
        "numeric_sum": sum(numeric_vals) if numeric_vals else 0,
        "bold_cells": bold_count, "colored_cells": colored,
    }

def get_headers(ws, header_row, max_col):
    headers = []
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        label = str(v).strip() if v not in [None, ""] else get_column_letter(c)
        headers.append((c, label))
    return headers

def get_data_matrix(ws, header_row, max_row, max_col):
    rows = []
    for r in range(header_row + 1, max_row + 1):
        row_data = {}
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            row_data[c] = "" if v is None else str(v)
        rows.append((r, row_data))
    return rows

def apply_filters(data_rows, active_filters):
    if not active_filters:
        return data_rows
    result = []
    for (r, row_data) in data_rows:
        match = True
        for col_idx, allowed_vals in active_filters.items():
            if allowed_vals and row_data.get(col_idx, "") not in allowed_vals:
                match = False
                break
        if match:
            result.append((r, row_data))
    return result

def worksheet_to_html(ws, max_row, max_col, header_row=4, search_term="", filtered_rows=None):
    merged, skip = merged_map(ws)
    search_lower  = search_term.strip().lower()
    headers       = get_headers(ws, header_row, max_col)

    html = ['<div class="table-wrap"><table class="excel-table">']
    html.append("<tr><th class='excel-corner'>#</th>")
    for (c, label) in headers:
        html.append(f"<th class='excel-col-header' title='{label}'>{label}</th>")
    html.append("</tr>")

    for r in range(1, header_row):
        html.append("<tr>")
        html.append(f"<th class='excel-row-header'>{r}</th>")
        for c in range(1, max_col + 1):
            if (r, c) in skip: continue
            cell  = ws.cell(row=r, column=c)
            value = "" if cell.value is None else str(cell.value)
            attrs = ""
            if (r, c) in merged:
                attrs += f' rowspan="{merged[(r,c)]["rowspan"]}" colspan="{merged[(r,c)]["colspan"]}"'
            html.append(f'<td{attrs} style="color:#9ca3af;background:#fafafa;font-style:italic;">{value}</td>')
        html.append("</tr>")

    rows_to_render = filtered_rows if filtered_rows is not None else [
        (r, {c: ("" if ws.cell(row=r, column=c).value is None else str(ws.cell(row=r, column=c).value))
             for c in range(1, max_col + 1)})
        for r in range(header_row + 1, max_row + 1)
    ]

    display_num = 1
    for (r, row_data) in rows_to_render:
        html.append("<tr>")
        html.append(f"<th class='excel-row-header'>{display_num}</th>")
        for c in range(1, max_col + 1):
            if (r, c) in skip: continue
            cell  = ws.cell(row=r, column=c)
            value = row_data.get(c, "")
            attrs = ""
            if (r, c) in merged:
                attrs += f' rowspan="{merged[(r,c)]["rowspan"]}" colspan="{merged[(r,c)]["colspan"]}"'
            extra = "search-match" if search_lower and search_lower in value.lower() else ""
            html.append(f'<td{attrs} class="{extra}" style="{cell_style(cell)}">{value}</td>')
        html.append("</tr>")
        display_num += 1

    html.append("</table></div>")
    return "".join(html)


# ─── COMPARISON ───────────────────────────────────────────────────────────────
@st.cache_resource(show_spinner="Loading workbook…")
def load_wb(path):
    return load_workbook(str(path), data_only=True, read_only=False)

def ws_to_dict(ws):
    data = {}
    max_row, max_col = 1, 1
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in [None, ""]:
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    for r in range(1, max_row+1):
        row_data = {}
        for c in range(1, max_col+1):
            v = ws.cell(row=r, column=c).value
            row_data[get_column_letter(c)] = "" if v is None else str(v)
        data[r] = row_data
    return data

def compare_sheets(ws_old, ws_new):
    old = ws_to_dict(ws_old)
    new = ws_to_dict(ws_new)
    added_rows, removed_rows, changed_rows = [], [], []
    all_rows = sorted(set(list(old.keys()) + list(new.keys())))
    for r in all_rows:
        if r in new and r not in old:
            added_rows.append({"row": r, "data": new[r]})
        elif r in old and r not in new:
            removed_rows.append({"row": r, "data": old[r]})
        elif r in old and r in new:
            diffs = []
            all_cols = sorted(set(list(old[r].keys()) + list(new[r].keys())))
            for col in all_cols:
                ov = old[r].get(col, "")
                nv = new[r].get(col, "")
                if ov != nv:
                    diffs.append({"col": col, "old": ov, "new": nv})
            if diffs:
                changed_rows.append({"row": r, "changes": diffs})
    return added_rows, removed_rows, changed_rows

def row_preview(row_data, max_cells=6):
    vals = [v for v in row_data.values() if v][:max_cells]
    return " · ".join(vals) if vals else "—"


# ─── PROJECT TRACKER ──────────────────────────────────────────────────────────
def load_projects():
    if TRACKER_FILE.exists():
        try:
            return json.loads(TRACKER_FILE.read_text())
        except Exception:
            return []
    return []

def save_projects(projects):
    TRACKER_FILE.write_text(json.dumps(projects, indent=2))

STATUS_OPTS   = ["Not Started", "In Progress", "On Hold", "Complete", "Cancelled"]
PRIORITY_OPTS = ["Low", "Medium", "High", "Critical"]
STATUS_CSS  = {"Complete":"st-complete","In Progress":"st-inprogress","On Hold":"st-onhold","Not Started":"st-notstarted","Cancelled":"st-cancelled"}
STATUS_CARD = {"Complete":"status-complete","In Progress":"status-inprogress","On Hold":"status-onhold","Not Started":"status-notstarted","Cancelled":"status-cancelled"}
PRIORITY_CSS = {"Critical":"pr-critical","High":"pr-high","Medium":"pr-medium","Low":"pr-low"}
PROG_CSS    = {"Complete":"c-complete","In Progress":"c-inprogress","On Hold":"c-onhold","Not Started":"c-notstarted","Cancelled":"c-cancelled"}


# ─── SUSTAINMENT UPDATES (Excel-backed) ───────────────────────────────────────
def get_or_create_updates_sheet():
    """Load the WIP xlsm and return (workbook, worksheet) for the updates sheet."""
    try:
        wb = load_workbook(str(WIP_UPDATES_FILE), keep_vba=True)
    except Exception as e:
        return None, str(e)

    if UPDATES_SHEET not in wb.sheetnames:
        ws_u = wb.create_sheet(UPDATES_SHEET)
        # Write header row with styling
        hdr_fill = PatternFill("solid", fgColor="1A4F8B")
        hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        thin = Side(style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for i, col_name in enumerate(UPDATE_COLS, 1):
            cell = ws_u.cell(row=1, column=i, value=col_name)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        # Set column widths
        widths = [20, 15, 15, 15, 15, 18, 10, 50]
        for i, w in enumerate(widths, 1):
            ws_u.column_dimensions[get_column_letter(i)].width = w
        try:
            wb.save(str(WIP_UPDATES_FILE))
        except Exception as e:
            return None, f"Could not save: {e}"
    return wb, None

def read_updates_from_excel():
    """Read all updates from the Dashboard Updates sheet."""
    try:
        wb = load_workbook(str(WIP_UPDATES_FILE), keep_vba=True, data_only=True)
        if UPDATES_SHEET not in wb.sheetnames:
            return []
        ws_u = wb[UPDATES_SHEET]
        updates = []
        for row in ws_u.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            updates.append({
                "timestamp": str(row[0]) if row[0] else "",
                "author":    str(row[1]) if row[1] else "",
                "wo":        str(row[2]) if row[2] else "",
                "customer":  str(row[3]) if row[3] else "",
                "pn":        str(row[4]) if row[4] else "",
                "type":      str(row[5]) if row[5] else "",
                "priority":  str(row[6]) if row[6] else "",
                "note":      str(row[7]) if row[7] else "",
            })
        return list(reversed(updates))  # newest first
    except Exception:
        return []

def append_update_to_excel(record):
    """Append one update row to the Dashboard Updates sheet in the WIP Excel."""
    try:
        wb = load_workbook(str(WIP_UPDATES_FILE), keep_vba=True)
        if UPDATES_SHEET not in wb.sheetnames:
            wb, err = get_or_create_updates_sheet()
            if err:
                return False, err
            wb = load_workbook(str(WIP_UPDATES_FILE), keep_vba=True)

        ws_u = wb[UPDATES_SHEET]
        next_row = ws_u.max_row + 1

        thin = Side(style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        font = Font(name="Calibri", size=11)

        values = [
            record["timestamp"], record["author"], record["wo"],
            record["customer"], record["pn"], record["type"],
            record["priority"], record["note"],
        ]
        for i, val in enumerate(values, 1):
            cell = ws_u.cell(row=next_row, column=i, value=val)
            cell.font = font
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(i == 8))
            # Alternate row shading
            if next_row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")

        wb.save(str(WIP_UPDATES_FILE))
        return True, None
    except PermissionError:
        return False, "The Excel file is currently open by someone else. Please close it and try again."
    except Exception as e:
        return False, str(e)


# ─── EDITABLE WIP FILE (full spreadsheet in-app) ──────────────────────────────
def load_wip_editable(sheet_name, header_row):
    """Read the WIP xlsm sheet into a DataFrame using header_row as column names."""
    wb = load_workbook(str(WIP_UPDATES_FILE), keep_vba=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return None, None, f"Sheet '{sheet_name}' not found."
    ws_e = wb[sheet_name]

    # Determine extent
    max_row, max_col = 1, 1
    for row in ws_e.iter_rows():
        for cell in row:
            if cell.value not in [None, ""]:
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)

    # Column names from header row
    columns = []
    seen = {}
    for c in range(1, max_col + 1):
        v = ws_e.cell(row=header_row, column=c).value
        name = str(v).strip() if v not in [None, ""] else get_column_letter(c)
        if name in seen:           # avoid duplicate column names
            seen[name] += 1
            name = f"{name} ({seen[name]})"
        else:
            seen[name] = 0
        columns.append(name)

    # Data rows after the header
    data = []
    for r in range(header_row + 1, max_row + 1):
        row_vals = []
        for c in range(1, max_col + 1):
            v = ws_e.cell(row=r, column=c).value
            row_vals.append("" if v is None else v)
        data.append(row_vals)

    df = pd.DataFrame(data, columns=columns)
    return df, (max_row, max_col), None

def save_wip_editable(df, sheet_name, header_row):
    """Write the edited DataFrame back into the WIP xlsm sheet, preserving macros."""
    try:
        wb = load_workbook(str(WIP_UPDATES_FILE), keep_vba=True)
        if sheet_name not in wb.sheetnames:
            return False, f"Sheet '{sheet_name}' not found."
        ws_e = wb[sheet_name]

        # Clear existing data rows (below the header)
        old_max = ws_e.max_row
        for r in range(header_row + 1, old_max + 1):
            for c in range(1, ws_e.max_column + 1):
                ws_e.cell(row=r, column=c).value = None

        # Write edited values
        font = Font(name="Calibri", size=11)
        for i, (_, row) in enumerate(df.iterrows()):
            excel_row = header_row + 1 + i
            for j, col_name in enumerate(df.columns, 1):
                val = row[col_name]
                if pd.isna(val) or val == "":
                    val = None
                cell = ws_e.cell(row=excel_row, column=j, value=val)
                cell.font = font

        wb.save(str(WIP_UPDATES_FILE))
        return True, None
    except PermissionError:
        return False, "The Excel file is open by someone else. Please close it and try again."
    except Exception as e:
        return False, str(e)


# ─── MASTHEAD ─────────────────────────────────────────────────────────────────
# Load the specific WIP file if it exists, otherwise newest in the folder
file = WIP_UPDATES_FILE if WIP_UPDATES_FILE.exists() else latest_excel(WIP_FOLDER)
now_str   = datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
file_name = file.name if file else "NO FILE FOUND"
mod_time  = datetime.fromtimestamp(os.path.getmtime(file)).strftime("%Y-%m-%d %H:%M") if file else "—"

st.markdown(f"""
<div class="masthead">
  <div class="masthead-icon">✈</div>
  <div>
    <div class="masthead-title">Sustainment WIP Dashboard</div>
    <div class="masthead-sub"><span class="status-dot"></span>AAIC Production · All Open Report</div>
  </div>
  <div class="masthead-meta">
    <div class="ts">⏱ {now_str}</div>
    <div class="fn">📂 {file_name} &nbsp;|&nbsp; Modified: {mod_time}</div>
  </div>
</div>
""", unsafe_allow_html=True)

if file is None:
    st.error("⚠ No Excel file found in the configured folder.")
    st.code(WIP_FOLDER)
    st.stop()


# ─── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div style="font-family:Rajdhani,sans-serif;font-size:1.25rem;font-weight:700;color:#1a4f8b;text-transform:uppercase;letter-spacing:0.1em;padding-bottom:10px;border-bottom:2px solid #e5e7eb;">⚙ Controls</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-sec">Worksheet</div>', unsafe_allow_html=True)
    workbook = load_wb(str(file))
    sheet    = st.selectbox("Select sheet", workbook.sheetnames, label_visibility="collapsed")
    ws       = workbook[sheet]
    det_rows, det_cols = detect_real_range(ws)

    st.markdown('<div class="sidebar-sec">Search / Filter</div>', unsafe_allow_html=True)
    search_term = st.text_input("Highlight cell value", placeholder="Type to search…", label_visibility="collapsed")

    st.markdown('<div class="sidebar-sec">Header & Range</div>', unsafe_allow_html=True)
    header_row = st.number_input("Header row number", min_value=1, max_value=50, value=4, step=1,
                                  help="Row in the Excel sheet that contains column names (e.g. row 4)")
    max_row = st.number_input("Max rows", min_value=1, max_value=10000, value=det_rows, step=10)
    max_col = st.number_input("Max cols", min_value=1, max_value=500,   value=det_cols, step=5)

    st.markdown('<div class="sidebar-sec">Actions</div>', unsafe_allow_html=True)
    if st.button("🔄  Refresh Data"):
        st.cache_resource.clear()
        st.rerun()

    st.markdown("---")
    st.markdown(f'<div style="font-size:0.7rem;color:#6b7280;line-height:1.7;word-break:break-all;">📁 {WIP_FOLDER}</div>', unsafe_allow_html=True)


# ─── KPI CARDS ────────────────────────────────────────────────────────────────
kpis      = collect_kpis(ws, int(max_row), int(max_col))
fill_pct  = round(kpis["non_empty"] / max(kpis["total_rows"]*kpis["total_cols"],1)*100, 1)
ns        = kpis["numeric_sum"]
num_fmt   = f"{ns:,.0f}" if abs(ns) < 1e9 else f"{ns/1e6:,.1f}M"

st.markdown(f"""
<div class="kpi-row">
  <div class="kpi-card blue">
    <div class="kpi-label">Active Rows</div>
    <div class="kpi-value">{kpis['total_rows']:,}</div>
    <div class="kpi-sub">{kpis['total_cols']} columns detected</div>
  </div>
  <div class="kpi-card green">
    <div class="kpi-label">Populated Cells</div>
    <div class="kpi-value">{kpis['non_empty']:,}</div>
    <div class="kpi-sub">{fill_pct}% fill rate</div>
  </div>
  <div class="kpi-card amber">
    <div class="kpi-label">Numeric Values</div>
    <div class="kpi-value">{kpis['numeric_count']:,}</div>
    <div class="kpi-sub">Sum: {num_fmt}</div>
  </div>
  <div class="kpi-card red">
    <div class="kpi-label">Highlighted Cells</div>
    <div class="kpi-value">{kpis['colored_cells']:,}</div>
    <div class="kpi-sub">{kpis['bold_cells']} bold cells</div>
  </div>
</div>
""", unsafe_allow_html=True)


# ─── TABS ─────────────────────────────────────────────────────────────────────
tab_wip, tab_diff, tab_tracker, tab_updates = st.tabs([
    "📋  WIP Viewer", "🔍  Day-Over-Day Comparison",
    "📌  Project Tracker", "📝  Sustainment Updates"
])


# ════════════════════════════════════════════════════════════
# TAB 1 — WIP VIEWER
# ════════════════════════════════════════════════════════════
with tab_wip:

    headers    = get_headers(ws, int(header_row), int(max_col))
    data_rows  = get_data_matrix(ws, int(header_row), int(max_row), int(max_col))

    col_unique = {}
    for (c, label) in headers:
        vals = sorted({row[c] for (_, row) in data_rows if row.get(c, "").strip()})
        if vals:
            col_unique[c] = (label, vals)

    active_filters = {}
    filter_cols_list = list(col_unique.items())

    st.markdown("""
    <div class="filter-panel">
      <div class="filter-panel-header">
        <span class="filter-panel-title">🔽 &nbsp;Column Filters</span>
        <span style="font-size:0.7rem;color:#a8c4e8;font-family:Calibri,sans-serif;">
          Select one or more values per column to narrow rows
        </span>
      </div>
      <div class="filter-panel-body">
    """, unsafe_allow_html=True)

    if "filter_clear_count" not in st.session_state:
        st.session_state["filter_clear_count"] = 0
    clear_count = st.session_state["filter_clear_count"]

    chunk_size = 4
    for i in range(0, len(filter_cols_list), chunk_size):
        chunk = filter_cols_list[i:i + chunk_size]
        ui_cols = st.columns(len(chunk))
        for ui_col, (col_idx, (label, unique_vals)) in zip(ui_cols, chunk):
            with ui_col:
                st.markdown(f'<div class="filter-col-label">▸ {label}</div>', unsafe_allow_html=True)
                selected = st.multiselect(
                    label, options=unique_vals, default=[],
                    key=f"filter_{sheet}_{col_idx}_v{clear_count}",
                    placeholder="All values", label_visibility="collapsed",
                )
                if selected:
                    active_filters[col_idx] = selected

    st.markdown('</div>', unsafe_allow_html=True)

    if active_filters:
        chips_html = ""
        for col_idx, vals in active_filters.items():
            col_label = col_unique[col_idx][0]
            for v in vals:
                chips_html += f'<span class="filter-chip"><span class="filter-chip-col">{col_label}:</span><span class="filter-chip-val">&nbsp;{v}</span></span>'
        st.markdown(f'<div class="filter-active-bar"><span style="font-size:0.72rem;font-weight:700;color:#1a4f8b;margin-right:4px;">ACTIVE FILTERS:</span>{chips_html}</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="filter-active-bar"><span class="filter-none-label">No filters active — showing all rows</span></div>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)

    cl1, _ = st.columns([1.4, 7])
    with cl1:
        if st.button("✖  Clear All Filters", key="clear_filters"):
            st.session_state["filter_clear_count"] += 1
            st.rerun()

    filtered = apply_filters(data_rows, active_filters)
    if search_term.strip():
        sl = search_term.strip().lower()
        filtered = [(r, rd) for (r, rd) in filtered if any(sl in v.lower() for v in rd.values())]

    total_data_rows = len(data_rows)
    shown_rows      = len(filtered)
    hidden_rows     = total_data_rows - shown_rows
    search_chip = f'<span class="filter-chip" style="background:#b45309;">⚡&nbsp;Search:&nbsp;{search_term.strip()}</span>' if search_term.strip() else ""
    hidden_note = f'<span style="font-size:0.78rem;color:#dc2626;font-weight:600;">{hidden_rows:,} rows hidden by filters</span>' if hidden_rows > 0 else ""

    st.markdown(f"""
    <div class="row-count-bar">
      <div><div class="row-count-number">{shown_rows:,}</div><div class="row-count-label">rows shown</div></div>
      <div class="row-count-divider"></div>
      <div><div class="row-count-number" style="color:#6b7280;">{total_data_rows:,}</div><div class="row-count-label">total rows</div></div>
      <div class="row-count-divider"></div>
      <div style="display:flex;align-items:center;gap:8px;flex:1;flex-wrap:wrap;">
        <span class="badge badge-new">{sheet}</span>{search_chip}{hidden_note}
      </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown(
        '<div class="panel" style="padding:0;overflow:hidden;">'
        + worksheet_to_html(ws, int(max_row), int(max_col), int(header_row), search_term, filtered)
        + '</div>', unsafe_allow_html=True
    )


# ════════════════════════════════════════════════════════════
# TAB 2 — DAY-OVER-DAY COMPARISON
# ════════════════════════════════════════════════════════════
with tab_diff:
    st.markdown('<div class="sec-head">Day-Over-Day WIP Comparison</div>', unsafe_allow_html=True)
    all_files = get_excel_files(WIP_FOLDER)
    if len(all_files) < 2:
        st.info("ℹ Only one Excel file found in the folder. A second file (from yesterday) is needed to compare.")
    else:
        col_a, col_b = st.columns(2)
        with col_a:
            file_labels = [f"{f.name}  ({datetime.fromtimestamp(os.path.getmtime(f)).strftime('%Y-%m-%d %H:%M')})" for f in all_files]
            new_idx = st.selectbox("Newer file (Today)", range(len(file_labels)), format_func=lambda i: file_labels[i], index=0)
        with col_b:
            old_idx = st.selectbox("Older file (Yesterday)", range(len(file_labels)), format_func=lambda i: file_labels[i], index=1)
        if new_idx == old_idx:
            st.warning("Please select two different files.")
        else:
            file_new = all_files[new_idx]
            file_old = all_files[old_idx]
            wb_new = load_wb(str(file_new))
            wb_old = load_wb(str(file_old))
            common_sheets = [s for s in wb_new.sheetnames if s in wb_old.sheetnames]
            new_only = [s for s in wb_new.sheetnames if s not in wb_old.sheetnames]
            old_only = [s for s in wb_old.sheetnames if s not in wb_new.sheetnames]
            compare_sheet = st.selectbox("Compare worksheet", common_sheets) if common_sheets else None
            if compare_sheet:
                added, removed, changed = compare_sheets(wb_old[compare_sheet], wb_new[compare_sheet])
                st.markdown(f"""
                <div class="kpi-row" style="margin-top:14px;">
                  <div class="kpi-card green"><div class="kpi-label">Rows Added</div><div class="kpi-value">{len(added)}</div><div class="kpi-sub">New entries today</div></div>
                  <div class="kpi-card red"><div class="kpi-label">Rows Removed</div><div class="kpi-value">{len(removed)}</div><div class="kpi-sub">Entries closed/removed</div></div>
                  <div class="kpi-card amber"><div class="kpi-label">Rows Modified</div><div class="kpi-value">{len(changed)}</div><div class="kpi-sub">Cell-level changes</div></div>
                  <div class="kpi-card blue"><div class="kpi-label">Net Change</div><div class="kpi-value">{len(added)-len(removed):+d}</div><div class="kpi-sub">Rows vs yesterday</div></div>
                </div>""", unsafe_allow_html=True)
                if new_only: st.markdown(f'<div style="margin:8px 0;font-size:0.8rem;color:#166534;">✚ New sheets today: {", ".join(new_only)}</div>', unsafe_allow_html=True)
                if old_only: st.markdown(f'<div style="margin:8px 0;font-size:0.8rem;color:#991b1b;">✖ Sheets removed: {", ".join(old_only)}</div>', unsafe_allow_html=True)
                if added:
                    with st.expander(f"✚  Added Rows ({len(added)})", expanded=True):
                        rows_html = "".join(f"<tr><td><b>{r['row']}</b></td><td><span class='badge badge-added'>Added</span></td><td>{row_preview(r['data'])}</td></tr>" for r in added)
                        st.markdown(f'<div style="overflow:auto;max-height:300px;"><table class="diff-table"><tr><th>Row #</th><th>Status</th><th>Preview</th></tr>{rows_html}</table></div>', unsafe_allow_html=True)
                if removed:
                    with st.expander(f"✖  Removed Rows ({len(removed)})", expanded=True):
                        rows_html = "".join(f"<tr><td><b>{r['row']}</b></td><td><span class='badge badge-removed'>Removed</span></td><td>{row_preview(r['data'])}</td></tr>" for r in removed)
                        st.markdown(f'<div style="overflow:auto;max-height:300px;"><table class="diff-table"><tr><th>Row #</th><th>Status</th><th>Preview</th></tr>{rows_html}</table></div>', unsafe_allow_html=True)
                if changed:
                    with st.expander(f"✎  Modified Rows ({len(changed)})", expanded=True):
                        rows_html = "".join(f"<tr><td><b>{r['row']}</b></td><td><span class='badge badge-changed'>Changed</span></td><td><b>{ch['col']}</b></td><td style='color:#991b1b;text-decoration:line-through'>{ch['old'] or '—'}</td><td style='color:#166534;font-weight:700'>{ch['new'] or '—'}</td></tr>" for r in changed for ch in r["changes"])
                        st.markdown(f'<div style="overflow:auto;max-height:400px;"><table class="diff-table"><tr><th>Row #</th><th>Status</th><th>Col</th><th>Old Value</th><th>New Value</th></tr>{rows_html}</table></div>', unsafe_allow_html=True)
                if not added and not removed and not changed:
                    st.success("✔ No differences found between the two files on this worksheet.")



# ════════════════════════════════════════════════════════════
# TAB 3 — PROJECT TRACKER
# ════════════════════════════════════════════════════════════
with tab_tracker:
    st.markdown('<div class="sec-head">Project Tracker</div>', unsafe_allow_html=True)
    projects = load_projects()
    with st.expander("➕  Add New Project", expanded=False):
        c1, c2 = st.columns(2)
        with c1:
            p_name   = st.text_input("Project Name *", key="p_name")
            p_owner  = st.text_input("Assigned To",    key="p_owner")
            p_status = st.selectbox("Status",    STATUS_OPTS,   key="p_status")
            p_prio   = st.selectbox("Priority",  PRIORITY_OPTS, key="p_prio", index=1)
        with c2:
            p_start  = st.date_input("Start Date", value=date.today(), key="p_start")
            p_due    = st.date_input("Due Date",   value=date.today(), key="p_due")
            p_pct    = st.slider("% Complete", 0, 100, 0, key="p_pct")
            p_desc   = st.text_area("Description / Notes", key="p_desc", height=100)
        if st.button("✔  Save Project"):
            if not p_name.strip():
                st.error("Project name is required.")
            else:
                projects.append({"id":str(uuid.uuid4()),"name":p_name.strip(),"owner":p_owner.strip(),"status":p_status,"priority":p_prio,"start":str(p_start),"due":str(p_due),"pct":p_pct,"desc":p_desc.strip(),"created":datetime.now().isoformat()})
                save_projects(projects)
                st.success(f"✔ Project '{p_name}' saved.")
                st.rerun()
    fc1, fc2, fc3 = st.columns([2, 2, 3])
    with fc1: f_status = st.selectbox("Filter by Status",   ["All"] + STATUS_OPTS,   key="f_status")
    with fc2: f_prio   = st.selectbox("Filter by Priority", ["All"] + PRIORITY_OPTS, key="f_prio")
    with fc3: f_search = st.text_input("Search projects…", placeholder="Name, owner, or keyword", key="f_search", label_visibility="collapsed")
    total_p = len(projects); done_p = sum(1 for p in projects if p["status"]=="Complete")
    active_p = sum(1 for p in projects if p["status"]=="In Progress")
    overdue_p = sum(1 for p in projects if p["status"] not in ["Complete","Cancelled"] and p.get("due","") < str(date.today()))
    st.markdown(f"""<div class="kpi-row" style="margin-top:10px;">
      <div class="kpi-card blue"><div class="kpi-label">Total Projects</div><div class="kpi-value">{total_p}</div></div>
      <div class="kpi-card green"><div class="kpi-label">Complete</div><div class="kpi-value">{done_p}</div></div>
      <div class="kpi-card amber"><div class="kpi-label">In Progress</div><div class="kpi-value">{active_p}</div></div>
      <div class="kpi-card red"><div class="kpi-label">Overdue</div><div class="kpi-value">{overdue_p}</div><div class="kpi-sub">Past due date</div></div>
    </div>""", unsafe_allow_html=True)
    filtered_p = [p for p in projects if (f_status=="All" or p["status"]==f_status) and (f_prio=="All" or p["priority"]==f_prio) and (not f_search.strip() or f_search.strip().lower() in (p["name"]+p.get("owner","")+p.get("desc","")).lower())]
    if not filtered_p:
        st.info("No projects match the current filters. Add one above.")
    else:
        for p in filtered_p:
            sc=STATUS_CSS.get(p["status"],"st-notstarted"); scc=STATUS_CARD.get(p["status"],"status-notstarted")
            pc=PRIORITY_CSS.get(p["priority"],"pr-medium"); pgc=PROG_CSS.get(p["status"],"c-notstarted"); pct=p.get("pct",0)
            is_overdue = p["status"] not in ["Complete","Cancelled"] and p.get("due","") < str(date.today())
            overdue_badge = '<span class="badge badge-removed" style="margin-left:6px;">OVERDUE</span>' if is_overdue else ""
            st.markdown(f"""<div class="proj-card {scc}"><div style="flex:1;">
              <div class="proj-title">{p['name']}</div>
              <div class="proj-meta">👤 {p.get('owner','—')} &nbsp;|&nbsp; 📅 {p.get('start','—')} → {p.get('due','—')} &nbsp;|&nbsp;
                <span class='badge {sc}'>{p['status']}</span><span class='badge {pc}' style='margin-left:5px;'>{p['priority']}</span>{overdue_badge}</div>
              {f'<div class="proj-desc">{p["desc"]}</div>' if p.get("desc") else ""}
              <div style="font-size:0.73rem;color:#6b7280;margin-bottom:6px;">{pct}% complete</div>
              <div class="prog-bar-wrap"><div class="prog-bar-fill {pgc}" style="width:{pct}%"></div></div>
            </div></div>""", unsafe_allow_html=True)
            with st.expander(f"⚙ Edit / Delete — {p['name']}", expanded=False):
                ec1, ec2 = st.columns(2)
                with ec1:
                    e_name=st.text_input("Name",value=p["name"],key=f"en_{p['id']}")
                    e_owner=st.text_input("Assigned",value=p.get("owner",""),key=f"eo_{p['id']}")
                    e_status=st.selectbox("Status",STATUS_OPTS,index=STATUS_OPTS.index(p["status"]),key=f"es_{p['id']}")
                    e_prio=st.selectbox("Priority",PRIORITY_OPTS,index=PRIORITY_OPTS.index(p["priority"]),key=f"ep_{p['id']}")
                with ec2:
                    e_start=st.date_input("Start",value=date.fromisoformat(p.get("start",str(date.today()))),key=f"est_{p['id']}")
                    e_due=st.date_input("Due",value=date.fromisoformat(p.get("due",str(date.today()))),key=f"edu_{p['id']}")
                    e_pct=st.slider("% Complete",0,100,p.get("pct",0),key=f"epc_{p['id']}")
                    e_desc=st.text_area("Notes",value=p.get("desc",""),key=f"ede_{p['id']}",height=80)
                b1, b2 = st.columns(2)
                with b1:
                    if st.button("💾  Update", key=f"upd_{p['id']}"):
                        for proj in projects:
                            if proj["id"]==p["id"]:
                                proj.update({"name":e_name.strip(),"owner":e_owner.strip(),"status":e_status,"priority":e_prio,"start":str(e_start),"due":str(e_due),"pct":e_pct,"desc":e_desc.strip()})
                        save_projects(projects); st.success("Updated."); st.rerun()
                with b2:
                    if st.button("🗑  Delete", key=f"del_{p['id']}"):
                        projects=[pr for pr in projects if pr["id"]!=p["id"]]
                        save_projects(projects); st.warning("Project deleted."); st.rerun()


# ════════════════════════════════════════════════════════════
# TAB 4 — SUSTAINMENT UPDATES  (live editable WIP spreadsheet)
# ════════════════════════════════════════════════════════════
with tab_updates:
    st.markdown('<div class="sec-head">Sustainment Updates — Live WIP List</div>', unsafe_allow_html=True)
    st.markdown(
        f'<div style="font-size:0.8rem;color:#374151;margin-bottom:10px;">'
        f'Edit directly below — changes save to <b>{WIP_UPDATES_FILE.name}</b>. '
        f'No need to open Excel.</div>',
        unsafe_allow_html=True
    )

    if not WIP_UPDATES_FILE.exists():
        st.error(f"⚠ Cannot reach the WIP file:\n{WIP_UPDATES_FILE}\n\nMake sure you are connected to the network drive.")
        st.stop()

    # ── Controls ─────────────────────────────────────────────
    try:
        wb_meta = load_workbook(str(WIP_UPDATES_FILE), keep_vba=True, read_only=True)
        wip_sheets = wb_meta.sheetnames
        wb_meta.close()
    except Exception as e:
        st.error(f"Could not open the WIP file: {e}")
        st.stop()

    ctrl1, ctrl2, ctrl3 = st.columns([3, 2, 2])
    with ctrl1:
        edit_sheet = st.selectbox("Worksheet to edit", wip_sheets, key="edit_sheet")
    with ctrl2:
        edit_header = st.number_input("Header row", min_value=1, max_value=50, value=4, step=1, key="edit_header",
                                       help="Row containing the column names (WORK_ORDER, PN, etc.)")
    with ctrl3:
        st.markdown("<div style='height:28px;'></div>", unsafe_allow_html=True)
        if st.button("🔄  Reload from Excel", key="reload_wip"):
            for k in list(st.session_state.keys()):
                if k.startswith("wip_editor_"):
                    del st.session_state[k]
            st.rerun()

    # ── Load the sheet into an editable grid ─────────────────
    df, extent, err = load_wip_editable(edit_sheet, int(edit_header))
    if err:
        st.error(err)
        st.stop()

    # KPI summary
    n_rows = len(df)
    n_cols = len(df.columns)
    n_filled = int((df.replace("", pd.NA).notna()).sum().sum())
    st.markdown(f"""<div class="kpi-row" style="margin-top:6px;">
      <div class="kpi-card blue"><div class="kpi-label">Rows</div><div class="kpi-value">{n_rows:,}</div><div class="kpi-sub">in {edit_sheet}</div></div>
      <div class="kpi-card green"><div class="kpi-label">Columns</div><div class="kpi-value">{n_cols}</div><div class="kpi-sub">from row {int(edit_header)}</div></div>
      <div class="kpi-card amber"><div class="kpi-label">Filled Cells</div><div class="kpi-value">{n_filled:,}</div></div>
      <div class="kpi-card red"><div class="kpi-label">Status</div><div class="kpi-value" style="font-size:1.1rem;">EDITABLE</div><div class="kpi-sub">type in any cell</div></div>
    </div>""", unsafe_allow_html=True)

    st.markdown('<div style="font-size:0.78rem;color:#6b7280;margin:6px 0;">💡 Tip: scroll to the bottom of the grid to add a new row. Click the ⋮ on a row to delete it.</div>', unsafe_allow_html=True)

    # ── Editable data grid ───────────────────────────────────
    edited_df = st.data_editor(
        df,
        use_container_width=True,
        num_rows="dynamic",
        height=560,
        key=f"wip_editor_{edit_sheet}_{int(edit_header)}",
    )

    # ── Save button ──────────────────────────────────────────
    sc1, sc2 = st.columns([1.5, 6])
    with sc1:
        if st.button("💾  Save Changes to Excel", key="save_wip"):
            ok, save_err = save_wip_editable(edited_df, edit_sheet, int(edit_header))
            if ok:
                st.success("✔ Saved to the WIP Excel file successfully!")
            else:
                st.error(f"❌ Could not save: {save_err}")
    with sc2:
        st.markdown(
            '<div style="font-size:0.75rem;color:#9ca3af;padding-top:8px;">'
            'Changes are written back to the network file and visible to everyone after they reload.</div>',
            unsafe_allow_html=True
        )
